import os
import django
from datetime import datetime, time

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'gf_backend.settings')
django.setup()

from employee.models import Employee, Role
from accounts.models import User
from company.models import Branch
from openpyxl import load_workbook

EXCEL_FILE_PATH = "emp_data.xlsx"


def parse_time(value):
    """
    Excel time can be:
    - datetime.time
    - string "13:00"
    - None
    """
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%H:%M").time()
    return None


def run():
    wb = load_workbook(EXCEL_FILE_PATH)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row = list(row)

        # 🔹 Column mapping (SAFE)
        name = row[0]
        email = row[1]
        joining_date = row[2]
        address = row[3]
        branch_id = row[4]
        role_id = row[5]
        # row[6] = additional_notes (ignored)
        base_salary = row[7]
        working_hours = row[8]
        phone = str(row[9]).strip()
        shift_in = row[10]
        shift_out = row[11]

        try:
            branch = Branch.objects.get(id=branch_id)
            role = Role.objects.get(id=role_id)
        except:
            print(f"❌ Invalid Branch/Role for {name}")
            continue

        employee, created = Employee.objects.get_or_create(
            phone=phone,
            defaults={
                "name": name,
                "email": email,
                "joining_date": joining_date,
                "address": address,
                "branch": branch,
                "role": role,
                "base_salary": base_salary,
                "working_hours": int(working_hours),
                "shift_in": parse_time(shift_in),
                "shift_out": parse_time(shift_out),
                "status": "active",
            }
        )

        if not employee.user:
            if User.objects.filter(username=phone).exists():
                print(f"⚠️ User exists for {phone}")
            else:
                user = User.objects.create_user(
                    username=phone,
                    password=phone,
                    role="employee",
                    branch=branch,
                    email=email or None
                )
                employee.user = user
                employee.save()
                print(f"✅ Employee + User created: {name}")
        else:
            print(f"ℹ️ Already linked: {name}")


if __name__ == "__main__":
    run()
